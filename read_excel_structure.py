#!/usr/bin/env python3
"""Temporary script to read Excel structure"""
import sys
try:
    import openpyxl
    wb = openpyxl.load_workbook('2025_Tax_Organizer_final_unprotected.xlsx')
    print("SHEET_NAMES:", ",".join(wb.sheetnames))
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
            if row_idx == 0:
                headers = [str(cell) if cell is not None else "" for cell in row]
                print(f"HEADERS_{sheet_name}:", "|".join(headers))
except ImportError:
    print("ERROR: openpyxl not installed", file=sys.stderr)
    sys.exit(1)
except Exception as e:
    print(f"ERROR: {e}", file=sys.stderr)
    sys.exit(1)
